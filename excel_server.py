#!/usr/bin/env python3
"""
Excel MCP Server - Provides tools for reading, writing, and manipulating Excel files.
Uses openpyxl, xlwings, and pandas for comprehensive Excel operations.
"""

import sys
import json
import os
import traceback
from typing import Any
from mcp.server import Server, NotificationOptions
from mcp.server.models import InitializationOptions
import mcp.server.stdio
import mcp.types as types

# Excel libraries
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import xlwings as xw

# Track open workbooks for resource management
_open_workbooks: dict = {}

class ExcelMCPServer:
    def __init__(self):
        self.server = Server("excel-server")
        self.setup_handlers()

    def setup_handlers(self):
        @self.server.list_tools()
        async def handle_list_tools() -> list[types.Tool]:
            return [
                types.Tool(
                    name="list_workbooks",
                    description="List all open Excel workbooks",
                    inputSchema={
                        "type": "object",
                        "properties": {},
                    },
                ),
                types.Tool(
                    name="open_workbook",
                    description="Open an Excel workbook file",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "visible": {
                                "type": "boolean",
                                "description": "Whether to make Excel visible (default: False)",
                                "default": False,
                            },
                        },
                        "required": ["filepath"],
                    },
                ),
                types.Tool(
                    name="create_workbook",
                    description="Create a new Excel workbook",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path where to save the new workbook",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the initial sheet (default: Sheet1)",
                                "default": "Sheet1",
                            },
                        },
                        "required": ["filepath"],
                    },
                ),
                types.Tool(
                    name="read_range",
                    description="Read data from a range of cells in a worksheet",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the worksheet",
                            },
                            "range_address": {
                                "type": "string",
                                "description": "Range address (e.g., 'A1:C10' or 'A1')",
                            },
                            "header": {
                                "type": "boolean",
                                "description": "Whether first row is a header",
                                "default": True,
                            },
                        },
                        "required": ["filepath", "sheet_name"],
                    },
                ),
                types.Tool(
                    name="write_range",
                    description="Write data to a range of cells in a worksheet",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the worksheet",
                            },
                            "start_cell": {
                                "type": "string",
                                "description": "Starting cell address (e.g., 'A1')",
                            },
                            "data": {
                                "type": "array",
                                "description": "2D array of data to write (list of lists)",
                                "items": {
                                    "type": "array",
                                    "items": {
                                        "type": ["string", "number", "boolean", "null"],
                                    },
                                },
                            },
                        },
                        "required": ["filepath", "sheet_name", "data"],
                    },
                ),
                types.Tool(
                    name="create_sheet",
                    description="Create a new worksheet in a workbook",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name for the new worksheet",
                            },
                        },
                        "required": ["filepath", "sheet_name"],
                    },
                ),
                types.Tool(
                    name="list_sheets",
                    description="List all worksheets in a workbook",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                        },
                        "required": ["filepath"],
                    },
                ),
                types.Tool(
                    name="get_sheet_info",
                    description="Get information about a worksheet (dimensions, column count, row count)",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the worksheet",
                            },
                        },
                        "required": ["filepath", "sheet_name"],
                    },
                ),
                types.Tool(
                    name="delete_sheet",
                    description="Delete a worksheet from a workbook",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the worksheet to delete",
                            },
                        },
                        "required": ["filepath", "sheet_name"],
                    },
                ),
                types.Tool(
                    name="run_macro",
                    description="Run a VBA macro in Excel using xlwings",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file with macros",
                            },
                            "macro_name": {
                                "type": "string",
                                "description": "Name of the macro to run",
                            },
                        },
                        "required": ["filepath", "macro_name"],
                    },
                ),
                types.Tool(
                    name="save_workbook",
                    description="Save changes to an open workbook",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                        },
                        "required": ["filepath"],
                    },
                ),
                types.Tool(
                    name="close_workbook",
                    description="Close a workbook to free resources",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                        },
                        "required": ["filepath"],
                    },
                ),
                types.Tool(
                    name="apply_formatting",
                    description="Apply formatting to a range of cells",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the worksheet",
                            },
                            "range_address": {
                                "type": "string",
                                "description": "Range address (e.g., 'A1:C10')",
                            },
                            "bold": {
                                "type": "boolean",
                                "description": "Make text bold",
                            },
                            "font_color": {
                                "type": "string",
                                "description": "Font color in hex (e.g., 'FF0000' for red)",
                            },
                            "fill_color": {
                                "type": "string",
                                "description": "Fill color in hex (e.g., 'FFFF00' for yellow)",
                            },
                            "font_size": {
                                "type": "number",
                                "description": "Font size",
                            },
                            "alignment": {
                                "type": "string",
                                "description": "Text alignment: 'left', 'center', 'right'",
                                "enum": ["left", "center", "right"],
                            },
                        },
                        "required": ["filepath", "sheet_name", "range_address"],
                    },
                ),
                types.Tool(
                    name="create_pivot_table",
                    description="Create a PivotTable from a data range (using pandas)",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "source_sheet": {
                                "type": "string",
                                "description": "Source worksheet name",
                            },
                            "pivot_sheet": {
                                "type": "string",
                                "description": "Destination worksheet name for the pivot table",
                            },
                            "rows": {
                                "type": "array",
                                "description": "Column names to use as row fields",
                                "items": {"type": "string"},
                            },
                            "values": {
                                "type": "array",
                                "description": "Column names to aggregate",
                                "items": {"type": "string"},
                            },
                            "aggfunc": {
                                "type": "string",
                                "description": "Aggregation function: 'sum', 'mean', 'count', 'min', 'max'",
                                "default": "sum",
                                "enum": ["sum", "mean", "count", "min", "max"],
                            },
                            "columns": {
                                "type": "array",
                                "description": "Column names to use as column fields (optional)",
                                "items": {"type": "string"},
                            },
                        },
                        "required": ["filepath", "source_sheet", "pivot_sheet", "rows", "values"],
                    },
                ),
                types.Tool(
                    name="create_chart",
                    description="Create a chart from worksheet data",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "sheet_name": {
                                "type": "string",
                                "description": "Name of the worksheet with data",
                            },
                            "chart_type": {
                                "type": "string",
                                "description": "Type of chart: 'bar', 'line', 'pie', 'scatter'",
                                "enum": ["bar", "line", "pie", "scatter"],
                            },
                            "title": {
                                "type": "string",
                                "description": "Chart title",
                            },
                            "categories": {
                                "type": "string",
                                "description": "Range for categories (e.g., 'A2:A10')",
                            },
                            "values": {
                                "type": "string",
                                "description": "Range for values (e.g., 'B2:B10')",
                            },
                            "output_cell": {
                                "type": "string",
                                "description": "Cell where to place the chart (e.g., 'D1')",
                                "default": "E1",
                            },
                        },
                        "required": ["filepath", "sheet_name", "chart_type", "title", "categories", "values"],
                    },
                ),
                types.Tool(
                    name="run_power_query",
                    description="Run a Power Query refresh using xlwings",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "filepath": {
                                "type": "string",
                                "description": "Full path to the Excel file",
                            },
                            "query_name": {
                                "type": "string",
                                "description": "Name of the Power Query to refresh (optional, refreshes all if omitted)",
                            },
                        },
                        "required": ["filepath"],
                    },
                ),
            ]

        @self.server.call_tool()
        async def handle_call_tool(name: str, arguments: dict[str, Any]) -> list[types.TextContent]:
            try:
                if name == "list_workbooks":
                    return await self._list_workbooks()
                elif name == "open_workbook":
                    return await self._open_workbook(arguments)
                elif name == "create_workbook":
                    return await self._create_workbook(arguments)
                elif name == "read_range":
                    return await self._read_range(arguments)
                elif name == "write_range":
                    return await self._write_range(arguments)
                elif name == "create_sheet":
                    return await self._create_sheet(arguments)
                elif name == "list_sheets":
                    return await self._list_sheets(arguments)
                elif name == "get_sheet_info":
                    return await self._get_sheet_info(arguments)
                elif name == "delete_sheet":
                    return await self._delete_sheet(arguments)
                elif name == "run_macro":
                    return await self._run_macro(arguments)
                elif name == "save_workbook":
                    return await self._save_workbook(arguments)
                elif name == "close_workbook":
                    return await self._close_workbook(arguments)
                elif name == "apply_formatting":
                    return await self._apply_formatting(arguments)
                elif name == "create_pivot_table":
                    return await self._create_pivot_table(arguments)
                elif name == "create_chart":
                    return await self._create_chart(arguments)
                elif name == "run_power_query":
                    return await self._run_power_query(arguments)
                else:
                    raise ValueError(f"Unknown tool: {name}")
            except Exception as e:
                return [types.TextContent(type="text", text=f"Error: {str(e)}\n{traceback.format_exc()}")]

    async def _list_workbooks(self) -> list[types.TextContent]:
        if not _open_workbooks:
            return [types.TextContent(type="text", text="[]")]
        wb_list = []
        for fp, wb in _open_workbooks.items():
            wb_list.append({
                "filepath": fp,
                "sheets": wb.sheetnames,
                "active_sheet": wb.active.title if wb.active else None,
            })
        return [types.TextContent(type="text", text=json.dumps(wb_list, indent=2))]

    async def _open_workbook(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        if not os.path.exists(filepath):
            return [types.TextContent(type="text", text=f"Error: File not found: {filepath}")]
        
        wb = load_workbook(filepath, data_only=True)
        _open_workbooks[filepath] = wb
        
        sheets = wb.sheetnames
        return [types.TextContent(type="text", text=json.dumps({
            "status": "opened",
            "filepath": filepath,
            "sheets": sheets,
            "active_sheet": wb.active.title,
        }, indent=2))]

    async def _create_workbook(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args.get("sheet_name", "Sheet1")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        wb.save(filepath)
        _open_workbooks[filepath] = wb
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "created",
            "filepath": filepath,
            "sheet_name": sheet_name,
        }, indent=2))]

    async def _read_range(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        header = args.get("header", True)
        
        wb = self._get_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")]
        
        ws = wb[sheet_name]
        
        if "range_address" in args and args["range_address"]:
            # Read specific range
            range_addr = args["range_address"]
            data = ws[range_addr]
            rows = []
            for row in data:
                rows.append([cell.value for cell in row])
            
            result = {
                "sheet": sheet_name,
                "range": range_addr,
                "data": rows,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
        else:
            # Read entire sheet
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                rows.append([cell.value for cell in row])
            
            result = {
                "sheet": sheet_name,
                "range": ws.dimensions,
                "data": rows,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
        
        return [types.TextContent(type="text", text=json.dumps(result, indent=2, default=str))]

    async def _write_range(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        start_cell = args.get("start_cell", "A1")
        data = args["data"]
        
        wb = self._get_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found.")]
        
        ws = wb[sheet_name]
        
        # Parse start cell coordinates
        col_letter = ""
        row_num = 0
        for i, ch in enumerate(start_cell):
            if ch.isalpha():
                col_letter += ch
            else:
                row_num = int(start_cell[i:])
                break
        
        start_col = openpyxl.utils.column_index_from_string(col_letter)
        
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                cell = ws.cell(row=row_num + i, column=start_col + j)
                cell.value = cell_value
        
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "written",
            "filepath": filepath,
            "sheet": sheet_name,
            "rows_written": len(data),
            "columns_written": len(data[0]) if data else 0,
        }, indent=2))]

    async def _create_sheet(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        
        wb = self._get_workbook(filepath)
        if sheet_name in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' already exists.")]
        
        ws = wb.create_sheet(title=sheet_name)
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "created",
            "sheet_name": sheet_name,
            "filepath": filepath,
        }, indent=2))]

    async def _list_sheets(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        wb = self._get_workbook(filepath)
        
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({
                "name": name,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            })
        
        return [types.TextContent(type="text", text=json.dumps({
            "filepath": filepath,
            "sheets": sheets,
        }, indent=2))]

    async def _get_sheet_info(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        
        wb = self._get_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found.")]
        
        ws = wb[sheet_name]
        
        return [types.TextContent(type="text", text=json.dumps({
            "filepath": filepath,
            "sheet": sheet_name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells": list(ws.merged_cells.ranges),
        }, indent=2, default=str))]

    async def _delete_sheet(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        
        wb = self._get_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found.")]
        
        if len(wb.sheetnames) <= 1:
            return [types.TextContent(type="text", text="Error: Cannot delete the only sheet in the workbook.")]
        
        del wb[sheet_name]
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "deleted",
            "sheet": sheet_name,
        }, indent=2))]

    async def _run_macro(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        macro_name = args["macro_name"]
        
        try:
            app = xw.App(visible=False)
            wb = app.books.open(filepath)
            macro = wb.macro(macro_name)
            result = macro()
            wb.save()
            app.quit()
            
            return [types.TextContent(type="text", text=json.dumps({
                "status": "macro_executed",
                "macro": macro_name,
                "result": str(result),
            }, indent=2))]
        except Exception as e:
            return [types.TextContent(type="text", text=f"Error running macro: {str(e)}")]

    async def _save_workbook(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        wb = self._get_workbook(filepath)
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "saved",
            "filepath": filepath,
        }, indent=2))]

    async def _close_workbook(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        
        if filepath in _open_workbooks:
            _open_workbooks[filepath].close()
            del _open_workbooks[filepath]
            return [types.TextContent(type="text", text=json.dumps({
                "status": "closed",
                "filepath": filepath,
            }, indent=2))]
        else:
            return [types.TextContent(type="text", text=f"Workbook not found in open list: {filepath}")]

    async def _apply_formatting(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        range_address = args["range_address"]
        
        wb = self._get_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found.")]
        
        ws = wb[sheet_name]
        cell_range = ws[range_address]
        
        for row in cell_range:
            for cell in row:
                if args.get("bold"):
                    cell.font = Font(bold=True, color=args.get("font_color"))
                if args.get("font_color"):
                    cell.font = Font(color=args.get("font_color"))
                if args.get("fill_color"):
                    cell.fill = PatternFill(start_color=args["fill_color"], end_color=args["fill_color"], fill_type="solid")
                if args.get("font_size"):
                    cell.font = Font(size=args["font_size"])
                if args.get("alignment"):
                    align_map = {"left": "left", "center": "center", "right": "right"}
                    cell.alignment = Alignment(horizontal=align_map.get(args["alignment"], "left"))
        
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "formatted",
            "range": range_address,
        }, indent=2))]

    async def _create_pivot_table(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        source_sheet = args["source_sheet"]
        pivot_sheet = args["pivot_sheet"]
        rows = args["rows"]
        values = args["values"]
        aggfunc = args.get("aggfunc", "sum")
        columns = args.get("columns", [])
        
        wb = self._get_workbook(filepath)
        if source_sheet not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{source_sheet}' not found.")]
        
        # Read data using pandas
        ws = wb[source_sheet]
        data = ws.values
        cols = next(data)
        df = pd.DataFrame(data, columns=cols)
        
        # Create pivot table
        agg_map = {
            "sum": "sum",
            "mean": "mean",
            "count": "count",
            "min": "min",
            "max": "max",
        }
        
        pivot_df = df.pivot_table(
            index=rows,
            values=values,
            columns=columns if columns else None,
            aggfunc=agg_map.get(aggfunc, "sum"),
            fill_value=0
        )
        
        # Write to new sheet
        if pivot_sheet in wb.sheetnames:
            del wb[pivot_sheet]
        
        ws_pivot = wb.create_sheet(title=pivot_sheet)
        
        # Write the pivot table data
        for r_idx, row in enumerate(pd.DataFrame(pivot_df.to_records()).values.tolist()):
            for c_idx, value in enumerate(row):
                ws_pivot.cell(row=r_idx + 1, column=c_idx + 1, value=value)
        
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "pivot_table_created",
            "sheet": pivot_sheet,
            "data_preview": str(pivot_df.head(10)),
        }, indent=2))]

    async def _create_chart(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        sheet_name = args["sheet_name"]
        chart_type = args["chart_type"]
        title = args["title"]
        categories = args["categories"]
        values = args["values"]
        output_cell = args.get("output_cell", "E1")
        
        wb = self._get_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return [types.TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found.")]
        
        ws = wb[sheet_name]
        
        # Create chart based on type
        chart_map = {
            "bar": openpyxl.chart.BarChart,
            "line": openpyxl.chart.LineChart,
            "pie": openpyxl.chart.PieChart,
            "scatter": openpyxl.chart.ScatterChart,
        }
        
        chart_class = chart_map.get(chart_type)
        if not chart_class:
            return [types.TextContent(type="text", text=f"Error: Unknown chart type: {chart_type}")]
        
        chart = chart_class()
        chart.title = title
        
        # Set up data references
        data_ref = openpyxl.chart.Reference(ws, range_string=f"'{sheet_name}'!{values}")
        cats_ref = openpyxl.chart.Reference(ws, range_string=f"'{sheet_name}'!{categories}")
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        if chart_type == "bar":
            chart.style = 10
        
        # Place chart
        ws.add_chart(chart, output_cell)
        wb.save(filepath)
        
        return [types.TextContent(type="text", text=json.dumps({
            "status": "chart_created",
            "type": chart_type,
            "title": title,
            "position": output_cell,
        }, indent=2))]

    async def _run_power_query(self, args: dict) -> list[types.TextContent]:
        filepath = os.path.abspath(args["filepath"])
        query_name = args.get("query_name")
        
        try:
            app = xw.App(visible=False)
            wb = app.books.open(filepath)
            
            if query_name:
                # Refresh specific query
                for conn in wb.api.Queries:
                    if conn.Name == query_name:
                        conn.Refresh()
                        break
                else:
                    app.quit()
                    return [types.TextContent(type="text", text=f"Error: Query '{query_name}' not found.")]
            else:
                # Refresh all
                wb.api.RefreshAll()
            
            wb.save()
            app.quit()
            
            return [types.TextContent(type="text", text=json.dumps({
                "status": "power_query_refreshed",
                "query": query_name or "all queries",
            }, indent=2))]
        except Exception as e:
            return [types.TextContent(type="text", text=f"Error refreshing Power Query: {str(e)}")]

    def _get_workbook(self, filepath: str) -> Workbook:
        """Get or load a workbook by filepath."""
        filepath = os.path.abspath(filepath)
        
        if filepath in _open_workbooks:
            return _open_workbooks[filepath]
        
        if os.path.exists(filepath):
            wb = load_workbook(filepath, data_only=True)
            _open_workbooks[filepath] = wb
            return wb
        
        raise FileNotFoundError(f"Workbook not found: {filepath}")

    async def run(self):
        async with mcp.server.stdio.stdio_server() as (read_stream, write_stream):
            await self.server.run(
                read_stream,
                write_stream,
                InitializationOptions(
                    server_name="excel-server",
                    server_version="1.0.0",
                    capabilities=self.server.get_capabilities(
                        notification_options=NotificationOptions(),
                        experimental_capabilities={},
                    ),
                ),
            )


if __name__ == "__main__":
    server = ExcelMCPServer()
    import asyncio
    asyncio.run(server.run())